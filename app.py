from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, send_file

from services.invoice_parser import parse_invoice_text, clean_invoice_text
from services.invoice_service import income_by_invoice_items, invoice_total
from services.notifications import get_notifications

from services.report_parser import (
    parse_iiko_report,
    extract_text_from_image,
    extract_text_from_pdf,
    clean_ocr_text
)

from services.writeoff_service import write_off_by_report_items
from services.audit_service import add_audit_log
from services.import_sales_service import create_sales_from_iiko_items

from models.audit import get_audit_logs
from models.recipes import get_all_recipes, get_recipes_grouped, create_recipe, get_recipe, delete_recipe, update_recipe_header, update_recipe_full
from models.locations import get_locations, get_location
from models.users import get_user_by_login

from models.inventory import (
    create_inventory_check,
    add_inventory_item,
    get_inventory_history,
    get_inventory_items
)

from models.products import (
    get_products,
    get_product,
    create_product,
    update_product,
    delete_product,
    low_stock_products
)

from models.sales import get_sales, create_sale
from models.expenses import get_expenses, create_expense

from models.stock import (
    get_stock_history,
    get_stock_balances,
    add_stock,
    write_off_stock,
    set_stock_balance,
)

from services.finance import get_finance_summary

from services.analytics import (
    dashboard_stats,
    top_products,
    sales_by_products,
    expenses_by_categories
)
from services.unit_converter import convert_quantity

from config import Config

import os
from io import BytesIO
from datetime import datetime
from werkzeug.utils import secure_filename
from models.products import get_products, get_product, create_product

app = Flask(__name__)
app.config.from_object(Config)


def login_required():
    return 'user' in session


def is_guest():
    return session.get('role') == 'guest'


def admin_required():
    if is_guest():
        flash('У гостя доступ только для просмотра', 'danger')
        return False
    return True


def current_location_id():
    return session.get('location_id')


def ensure_uploads_folder():
    if not os.path.exists('uploads'):
        os.makedirs('uploads')


def normalize_uploaded_filename(file):
    filename = secure_filename(file.filename)

    if '.' not in filename:
        if file.content_type == 'application/pdf':
            filename += '.pdf'
        elif file.content_type.startswith('image/'):
            filename += '.jpg'

    return filename




def report_filename(report_type, fmt):
    names = {
        'stock': 'ostatki_sklada',
        'movements': 'dvizhenie_sklada',
        'sales': 'prodazhi',
        'expenses': 'rashody',
        'inventory': 'revizii',
        'low_stock': 'nizkie_ostatki',
        'top_products': 'top_tovarov',
        'writeoffs': 'spisaniya_tovarov',
    }
    date_part = datetime.now().strftime('%Y-%m-%d_%H-%M')
    return f"{names.get(report_type, 'otchet')}_{date_part}.{fmt}"


def build_report_data(report_type, location_id):
    """Готовит данные для скачивания PDF/Excel отчетов."""
    if report_type == 'stock':
        rows = get_stock_balances(location_id)
        headers = ['Товар', 'Категория', 'Остаток', 'Ед.', 'Закуп. цена', 'Сумма остатка', 'Статус']
        data = []
        for row in rows:
            qty = float(row.get('quantity') or 0)
            price = float(row.get('purchase_price') or 0)
            min_qty = float(row.get('min_quantity') or 0)
            status = 'Нет в наличии' if qty <= 0 else ('Низкий остаток' if qty <= min_qty else 'Норма')
            data.append([row.get('name'), row.get('category'), qty, row.get('unit'), price, round(qty * price, 2), status])
        return 'Остатки на складе', headers, data

    if report_type == 'movements':
        rows = get_stock_history(location_id=location_id)
        headers = ['Дата', 'Тип операции', 'Товар', 'Количество', 'Комментарий']
        data = [[r.get('created_at'), 'Приход' if r.get('movement_type') == 'income' else 'Списание', r.get('product_name'), r.get('quantity'), r.get('comment')] for r in rows]
        return 'Движение товаров', headers, data

    if report_type == 'writeoffs':
        rows = [r for r in get_stock_history(location_id=location_id) if r.get('movement_type') == 'write_off']
        headers = ['Дата', 'Товар', 'Количество', 'Комментарий']
        data = [[r.get('created_at'), r.get('product_name'), r.get('quantity'), r.get('comment')] for r in rows]
        return 'Списания товаров', headers, data

    if report_type == 'sales':
        rows = get_sales(location_id)
        headers = ['№ продажи', 'Дата', 'Товар', 'Количество', 'Цена', 'Сумма']
        data = [[r.get('id'), r.get('created_at'), r.get('product_name'), r.get('quantity'), r.get('price'), r.get('total')] for r in rows]
        return 'Анализ продаж', headers, data

    if report_type == 'top_products':
        rows = top_products(location_id)
        headers = ['Товар', 'Количество', 'Выручка']
        data = [[r.get('name'), r.get('quantity'), r.get('total')] for r in rows]
        return 'Топ товаров', headers, data

    if report_type == 'expenses':
        rows = get_expenses(location_id)
        headers = ['Дата', 'Категория', 'Наименование', 'Сумма', 'Комментарий']
        data = [[r.get('created_at'), r.get('category'), r.get('title'), r.get('amount'), r.get('comment')] for r in rows]
        return 'Расходы', headers, data

    if report_type == 'inventory':
        rows = get_inventory_history(location_id)
        headers = ['№ ревизии', 'Дата', 'Склад', 'Пользователь', 'Статус']
        data = [[r.get('id'), r.get('created_at'), f"Склад №{r.get('location_id')}", r.get('username'), 'Завершена'] for r in rows]
        return 'История ревизий', headers, data

    if report_type == 'low_stock':
        rows = low_stock_products(location_id)
        headers = ['Товар', 'Категория', 'Остаток', 'Мин. остаток', 'Ед.']
        data = [[r.get('name'), r.get('category'), r.get('quantity'), r.get('min_quantity'), r.get('unit')] for r in rows]
        return 'Товары с низким остатком', headers, data

    return 'Отчет', ['Сообщение'], [['Нет данных для выбранного отчета']]


def create_excel_file(title, headers, data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Отчет'
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='FBE9D8')
    thin = Side(style='thin', color='E8D6C8')
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        cell.alignment = Alignment(horizontal='center')

    for row in data:
        ws.append(row)

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical='top')

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 3, 14), 38)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_pdf_file(title, headers, data):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = 'Helvetica'
    candidates = [
        'C:/Windows/Fonts/arial.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    for candidate in candidates:
        if os.path.exists(candidate):
            pdfmetrics.registerFont(TTFont('SmartCafeFont', candidate))
            font_name = 'SmartCafeFont'
            break

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    styles['Title'].fontName = font_name
    styles['Normal'].fontName = font_name

    table_data = [headers] + [[str(value or '') for value in row] for row in data]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FBE9D8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#5F2B0C')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E8D6C8')),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF8F2')]),
    ]))

    elements = [Paragraph(title, styles['Title']), Spacer(1, 12), table]
    doc.build(elements)
    output.seek(0)
    return output


def prepare_chart_items(rows, label_key, value_key):
    colors = ['#c86b2d', '#f0a35e', '#7b4a2d', '#d98d48', '#8f5d3d', '#efc18b']
    total = sum(float(row.get(value_key) or 0) for row in rows)
    items = []
    start = 0

    for index, row in enumerate(rows[:6]):
        value = float(row.get(value_key) or 0)
        percent = round((value / total * 100), 1) if total else 0
        end = start + percent
        items.append({
            'label': row.get(label_key),
            'value': round(value, 2),
            'percent': percent,
            'color': colors[index % len(colors)],
            'start': start,
            'end': end
        })
        start = end

    if not items:
        items.append({
            'label': 'Нет данных',
            'value': 0,
            'percent': 100,
            'color': '#e8d9ca',
            'start': 0,
            'end': 100
        })

    gradient = ', '.join([
        f"{item['color']} {item['start']}% {item['end']}%"
        for item in items
    ])

    return {
        'items': items,
        'gradient': gradient
    }


@app.before_request
def enforce_auth_and_location():
    """Сначала вход, затем обязательный выбор склада, только потом остальные вкладки."""
    public_endpoints = {
        'login',
        'logout',
        'select_location',
        'change_location',
        'static',
        'uploaded_file'
    }

    endpoint = request.endpoint

    if endpoint is None or endpoint in public_endpoints:
        return None

    if not login_required():
        return redirect(url_for('login'))

    if not session.get('location_id'):
        return redirect(url_for('select_location'))

    return None


@app.route('/', methods=['GET', 'POST'])
def login():

    if login_required():
        if session.get('location_id'):
            return redirect(url_for('dashboard'))
        return redirect(url_for('select_location'))

    if request.method == 'POST':

        user = get_user_by_login(
            request.form['username'],
            request.form['password']
        )

        if user:
            session['user'] = user['username']
            session['role'] = user['role']

            return redirect(url_for('select_location'))

        flash('Неверный логин или пароль', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():

    session.clear()
    return redirect(url_for('login'))


@app.route('/select-location', methods=['GET', 'POST'])
def select_location():

    if not login_required():
        return redirect(url_for('login'))

    if request.method == 'POST':

        location_id = request.form['location_id']
        location = get_location(location_id)

        location_map = {
            1: 'Мамыр',
            2: 'Розыбакиев',
            3: 'Абылайхана'
        }
        session['location_id'] = location['id']
        session['location_name'] = location_map.get(location['id'], location['name'])

        return redirect(url_for('dashboard'))

    return render_template(
        'select_location.html',
        locations=get_locations()
    )


@app.route('/change-location')
def change_location():

    if not login_required():
        return redirect(url_for('login'))

    session.pop('location_id', None)
    session.pop('location_name', None)

    return redirect(url_for('select_location'))


@app.route('/dashboard')
def dashboard():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()

    sales_top = top_products(location_id)
    expense_top = expenses_by_categories(location_id)

    return render_template(
        'dashboard.html',
        stats=dashboard_stats(location_id),
        finance=get_finance_summary(location_id),
        low_stock=low_stock_products(location_id),
        top_products=sales_top,
        top_expenses=expense_top,
        sales_chart=prepare_chart_items(sales_top, 'name', 'total'),
        expenses_chart=prepare_chart_items(expense_top, 'category', 'total'),
        recent_logs=get_audit_logs(location_id)[:8],
        recent_movements=get_stock_history(location_id=location_id)[:8],
        notifications=[]
    )


@app.route('/products', methods=['GET', 'POST'])
def products():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()

    if request.method == 'POST':

        if not admin_required():
            return redirect(url_for('products'))

        create_product(
            request.form['name'],
            request.form['category'],
            request.form['unit'],
            request.form['purchase_price'],
            request.form['sale_price'],
            request.form['quantity'],
            request.form['min_quantity'],
            location_id
        )

        add_audit_log(
            session.get('user'),
            location_id,
            'Создание товара',
            request.form['name'],
            'Товар добавлен вручную'
        )

        flash('Товар добавлен', 'success')

        return redirect(url_for('products'))

    search = request.args.get('search', '')

    return render_template(
        'products.html',
        products=get_products(search, location_id),
        search=search
    )


@app.route('/products/<int:product_id>/edit', methods=['GET', 'POST'])
def edit_product(product_id):

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()
    product = get_product(product_id, location_id)

    if request.method == 'POST':

        if not admin_required():
            return redirect(url_for('products'))

        update_product(
            product_id,
            request.form['name'],
            request.form['category'],
            request.form['unit'],
            request.form['purchase_price'],
            request.form.get('sale_price', 0),
            request.form['quantity'],
            request.form['min_quantity'],
            location_id
        )

        set_stock_balance(
            product_id,
            request.form['quantity'],
            location_id
        )

        add_audit_log(
            session.get('user'),
            location_id,
            'Изменение товара',
            request.form['name'],
            'Данные товара и остаток на складе обновлены'
        )

        flash('Товар обновлен', 'success')

        return redirect(url_for('stock'))

    return render_template(
        'edit_product.html',
        product=product
    )

@app.route('/products/<int:product_id>/delete')
def remove_product(product_id):

    if not login_required():
        return redirect(url_for('login'))

    if not admin_required():
        return redirect(url_for('products'))

    location_id = current_location_id()

    product = get_product(product_id, location_id)

    delete_product(product_id, location_id)

    add_audit_log(
        session.get('user'),
        location_id,
        'Удаление товара',
        product['name'] if product else 'Товар',
        'Товар удален из системы'
    )

    flash('Товар удален', 'success')

    return redirect(url_for('products'))


@app.route('/sales', methods=['GET', 'POST'])
def sales():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()

    if request.method == 'POST':

        if not admin_required():
            return redirect(url_for('sales'))

        product_id = request.form['product_id']
        sold_product = get_product(product_id, location_id)
        dish_name = sold_product['name']

        quantity = float(request.form['quantity'])
        price = float(request.form['price'])
        total = quantity * price

        if total > 1000000:
            flash('Сумма продажи не должна превышать 1 000 000 ₸', 'danger')
            return redirect(url_for('sales'))

        create_sale(
            product_id,
            quantity,
            price,
            location_id
        )

        recipe_items = get_recipe(dish_name)

        for item in recipe_items:
            ingredient_quantity = float(item['amount']) * quantity

            write_off_stock(
                item['product_id'],
                ingredient_quantity,
                f'Автосписание по продаже: {dish_name} x {quantity}',
                location_id
            )

        add_audit_log(
            session.get('user'),
            location_id,
            'Продажа',
            dish_name,
            'Продажа добавлена вручную с автосписанием ингредиентов'
        )

        flash('Продажа добавлена, ингредиенты списаны по техкарте', 'success')

        return redirect(url_for('sales'))

    return render_template(
        'sales.html',
        sales=get_sales(location_id),
        products=get_products(location_id=location_id)
    )

@app.route('/expenses', methods=['GET', 'POST'])
def expenses():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()

    if request.method == 'POST':

        if not admin_required():
            return redirect(url_for('expenses'))

        create_expense(
            request.form['title'],
            request.form['category'],
            request.form['amount'],
            request.form.get('comment', ''),
            location_id
        )

        add_audit_log(
            session.get('user'),
            location_id,
            'Расход',
            request.form['title'],
            'Расход добавлен вручную'
        )

        flash('Расход добавлен', 'success')

        return redirect(url_for('expenses'))

    return render_template(
        'expenses.html',
        expenses=get_expenses(location_id)
    )

@app.route('/stock', methods=['GET', 'POST'])
def stock():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()

    if request.method == 'POST':

        if not admin_required():
            return redirect(url_for('stock'))

        operation_type = request.form.get('operation_type')

        if operation_type == 'add_product':
            name = request.form.get('name', '').strip()
            quantity = request.form.get('quantity', '0')
            unit = request.form.get('unit', 'шт')
            purchase_price = request.form.get('purchase_price', '0')
            min_quantity = request.form.get('min_quantity', '0')

            if not name.replace(' ', '').isalpha():
                flash('Наименование товара должно содержать только буквы', 'danger')
                return redirect(url_for('stock'))

            existing_products = get_products(name, location_id)
            existing_product = None

            for item in existing_products:
                if item['name'].strip().lower() == name.lower():
                    existing_product = item
                    break

            if existing_product:
                product_id = existing_product['id']
                stock_quantity = convert_quantity(quantity, unit, existing_product['unit'])
            else:
                create_product(
                    name,
                    'Сырье',
                    unit,
                    purchase_price,
                    0,
                    0,
                    min_quantity,
                    location_id
                )

                product_id = get_products(name, location_id)[0]['id']
                stock_quantity = quantity

            add_stock(
                product_id,
                stock_quantity,
                'Ручное добавление товара',
                location_id
            )

            add_audit_log(
                session.get('user'),
                location_id,
                'Добавление товара вручную',
                name,
                'Товар добавлен вручную, так как не все ингредиенты поступают через накладную'
            )

            flash('Товар добавлен вручную', 'success')
            return redirect(url_for('stock'))

        if operation_type == 'edit_product':
            product_id = request.form.get('product_id')
            name = request.form.get('name', '').strip()
            category = request.form.get('category', '').strip()
            unit = request.form.get('unit', 'шт')
            purchase_price = request.form.get('purchase_price', '0')
            min_quantity = request.form.get('min_quantity', '0')
            quantity = request.form.get('quantity', '0')

            update_product(
                product_id,
                name,
                category,
                unit,
                purchase_price,
                0,
                0,
                min_quantity,
                location_id
            )

            set_stock_balance(
                product_id,
                quantity,
                location_id
            )

            add_audit_log(
                session.get('user'),
                location_id,
                'Редактирование товара',
                name,
                'Данные товара и текущий остаток обновлены через склад'
            )

            flash('Товар обновлен', 'success')
            return redirect(url_for('stock'))

        if operation_type == 'delete_product':
            product_id = request.form.get('product_id')
            product = get_product(product_id, location_id)

            if product:
                delete_product(product_id, location_id)

                add_audit_log(
                    session.get('user'),
                    location_id,
                    'Удаление товара',
                    product['name'],
                    'Товар удален из склада'
                )

                flash('Товар удален', 'success')
            else:
                flash('Товар не найден', 'danger')

            return redirect(url_for('stock'))

        product_id = request.form.get('product_id')
        quantity = request.form.get('quantity')
        quantity_unit = request.form.get('quantity_unit', '')
        action = request.form.get('action')
        comment = request.form.get('comment', '').strip()

        product = get_product(product_id, location_id)

        if not product:
            flash('Товар не найден', 'danger')
            return redirect(url_for('stock'))

        stock_quantity = convert_quantity(
            quantity,
            quantity_unit or product['unit'],
            product['unit']
        )

        if action == 'transfer':
            target_location_id = request.form.get('target_location_id')

            write_off_stock(
                product_id,
                stock_quantity,
                f'Перемещение на другой склад: {comment}',
                location_id
            )

            target_products = get_products(product['name'], target_location_id)
            target_product = None

            for item in target_products:
                if item['name'].strip().lower() == product['name'].strip().lower() and item['unit'] == product['unit']:
                    target_product = item
                    break

            if target_product:
                target_product_id = target_product['id']
            else:
                create_product(
                    product['name'],
                    product['category'],
                    product['unit'],
                    product['purchase_price'],
                    product['sale_price'],
                    0,
                    product['min_quantity'],
                    target_location_id
                )

                target_product_id = get_products(product['name'], target_location_id)[0]['id']

            add_stock(
                target_product_id,
                stock_quantity,
                f'Перемещение со склада: {comment}',
                target_location_id
            )

            add_audit_log(
                session.get('user'),
                location_id,
                'Перемещение товара',
                product['name'],
                comment or 'Перемещение товара между складами'
            )

            flash('Перемещение выполнено', 'success')

        elif action == 'writeoff':

            if not comment:
                flash('Причина списания обязательна для заполнения', 'danger')
                return redirect(url_for('stock'))

            write_off_stock(
                product_id,
                stock_quantity,
                comment,
                location_id
            )

            add_audit_log(
                session.get('user'),
                location_id,
                'Списание товара',
                product['name'],
                comment
            )

            flash('Списание выполнено', 'success')

        return redirect(url_for('stock'))

    return render_template(
        'stock.html',
        products=get_stock_balances(location_id)
    )

@app.route('/movements')
def movements():

    if not login_required():
        return redirect(url_for('login'))

    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    history = get_stock_history(
        search,
        date_from,
        date_to,
        current_location_id()
    )

    return render_template(
        'movements.html',
        history=history,
        search=search,
        date_from=date_from,
        date_to=date_to
    )


# Вкладка «Финансы» убрана из интерфейса. Финансовые расчеты используются только на главной.


@app.route('/analytics')
def analytics():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()

    sales_data = sales_by_products(location_id)
    expense_data = expenses_by_categories(location_id)

    return render_template(
        'analytics.html',
        finance=get_finance_summary(location_id),
        sales_labels=[row['name'] for row in sales_data],
        sales_values=[float(row['total']) for row in sales_data],
        expense_labels=[row['category'] for row in expense_data],
        expense_values=[float(row['total']) for row in expense_data]
    )




@app.route('/download-report/<report_type>/<fmt>')
def download_report(report_type, fmt):

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()
    title, headers, data = build_report_data(report_type, location_id)

    if fmt == 'excel':
        output = create_excel_file(title, headers, data)
        return send_file(
            output,
            as_attachment=True,
            download_name=report_filename(report_type, 'xlsx'),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    if fmt == 'pdf':
        output = create_pdf_file(title, headers, data)
        return send_file(
            output,
            as_attachment=True,
            download_name=report_filename(report_type, 'pdf'),
            mimetype='application/pdf'
        )

    flash('Неверный формат отчета', 'danger')
    return redirect(request.referrer or url_for('dashboard'))


# Вкладка «Отчеты» убрана из интерфейса. Скачивание PDF/Excel работает через /download-report.


@app.route('/import-report', methods=['GET', 'POST'])
def import_report():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()

    filename = None
    image_url = None
    scanned_text = ''
    parsed_items = []
    writeoff_results = []
    sales_results = []

    if request.method == 'POST':

        if not admin_required():
            return redirect(url_for('import_report'))

        report_text = request.form.get('report_text')
        action = request.form.get('action')
        file = request.files.get('report')

        if report_text:

            scanned_text = clean_ocr_text(report_text)
            parsed_items = parse_iiko_report(scanned_text)

            if action == 'writeoff':

                # 1) Добавляем готовые позиции из отчета iiko в продажи
                sales_results = create_sales_from_iiko_items(
                    parsed_items,
                    location_id
                )

                # 2) Списываем ингредиенты по техкартам со склада
                # Остатки могут уходить в минус — это разрешено по ТЗ.
                writeoff_results = write_off_by_report_items(
                    parsed_items,
                    location_id
                )

                add_audit_log(
                    session.get('user'),
                    location_id,
                    'Импорт отчета iiko',
                    'Отчет iiko',
                    'Продажи добавлены, ингредиенты списаны по техкартам'
                )

                flash('Импорт выполнен: продажи добавлены, ингредиенты списаны', 'success')

        elif file and file.filename:

            filename = normalize_uploaded_filename(file)
            ensure_uploads_folder()

            file_path = os.path.join('uploads', filename)
            file.save(file_path)

            if filename.lower().endswith('.pdf'):
                scanned_text = clean_ocr_text(
                    extract_text_from_pdf(file_path)
                )
            else:
                scanned_text = clean_ocr_text(
                    extract_text_from_image(file_path)
                )

            parsed_items = parse_iiko_report(scanned_text)

            image_url = url_for(
                'uploaded_file',
                filename=filename
            )

            flash(
                'Текст распознан. При необходимости исправьте вручную.',
                'success'
            )

    return render_template(
        'import_report.html',
        filename=filename,
        image_url=image_url,
        scanned_text=scanned_text,
        parsed_items=parsed_items,
        writeoff_results=writeoff_results,
        sales_results=sales_results
    )


@app.route('/import-invoice', methods=['GET', 'POST'])
def import_invoice():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()

    scanned_text = ''
    invoice_items = []
    income_results = []

    if request.method == 'POST':

        if not admin_required():
            return redirect(url_for('import_invoice'))

        invoice_text = request.form.get('invoice_text')
        action = request.form.get('action')
        file = request.files.get('invoice')

        if invoice_text:

            scanned_text = clean_invoice_text(invoice_text)
            invoice_items = parse_invoice_text(scanned_text)

            if action == 'income':

                income_results = income_by_invoice_items(
                    invoice_items,
                    location_id
                )

                total_amount = invoice_total(invoice_items)

                if total_amount > 0:
                    create_expense(
                        'Закупка по накладной',
                        'Закуп товара',
                        total_amount,
                        'Автоматически создано при импорте накладной',
                        location_id
                    )

                add_audit_log(
                    session.get('user'),
                    location_id,
                    'Оприходование',
                    'Накладная',
                    f'Товары оприходованы. Расход создан на сумму {total_amount} ₸'
                )

                flash(f'Оприходование выполнено. Расход добавлен: {total_amount} ₸', 'success')

        elif file and file.filename:

            filename = normalize_uploaded_filename(file)
            ensure_uploads_folder()

            file_path = os.path.join('uploads', filename)
            file.save(file_path)

            if filename.lower().endswith('.pdf'):
                raw_text = extract_text_from_pdf(file_path)
            else:
                raw_text = extract_text_from_image(file_path)

            scanned_text = clean_invoice_text(raw_text)

    return render_template(
        'import_invoice.html',
        scanned_text=scanned_text,
        invoice_items=invoice_items,
        income_results=income_results
    )


@app.route('/uploads/<filename>')
def uploaded_file(filename):

    return send_from_directory(
        'uploads',
        filename
    )


@app.route('/recipes', methods=['GET', 'POST'])
def recipes():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()

    if request.method == 'POST':

        if not admin_required():
            return redirect(url_for('recipes'))

        operation_type = request.form.get('operation_type')

        if operation_type == 'delete_recipe':
            dish_name = request.form.get('dish_name')

            delete_recipe(dish_name)

            add_audit_log(
                session.get('user'),
                location_id,
                'Удаление техкарты',
                dish_name,
                'Техкарта удалена'
            )
        if operation_type == 'edit_recipe':
            old_dish_name = request.form.get('old_dish_name')
            dish_name = request.form.get('dish_name')
            output_quantity = request.form.get('output_quantity', 1)
            output_unit = request.form.get('output_unit', 'ml')

            product_ids = request.form.getlist('product_id')
            amounts = request.form.getlist('amount')
            units = request.form.getlist('unit')

            valid_items = [
                item for item in zip(product_ids, amounts, units)
                if item[0] and item[1] and item[2]
            ]

            if not valid_items:
                flash('Нельзя сохранить техкарту без ингредиентов', 'danger')
                return redirect(url_for('recipes'))

            update_recipe_full(
                old_dish_name,
                dish_name,
                output_quantity,
                output_unit,
                product_ids,
                amounts,
                units
            )

            add_audit_log(
                session.get('user'),
                location_id,
                'Редактирование техкарты',
                dish_name,
                'Техкарта и ингредиенты обновлены'
            )

            flash('Техкарта обновлена', 'success')
            return redirect(url_for('recipes'))

        dish_name = request.form['dish_name']
        delete_recipe(dish_name)
        output_quantity = request.form.get('output_quantity', 1)
        output_unit = request.form.get('output_unit', 'порция')

        product_ids = request.form.getlist('product_id')
        amounts = request.form.getlist('amount')
        units = request.form.getlist('unit')

        added_count = 0

        for product_id, amount, unit in zip(product_ids, amounts, units):
            if not product_id or not amount or not unit:
                continue

            create_recipe(
                dish_name,
                product_id,
                amount,
                unit,
                None,
                output_quantity,
                output_unit
            )

            added_count += 1

        add_audit_log(
            session.get('user'),
            location_id,
            'Создание техкарты',
            dish_name,
            f'Добавлено ингредиентов: {added_count}'
        )

        flash('Техкарта сохранена', 'success')
        return redirect(url_for('recipes'))

    return render_template(
        'recipes.html',
        recipes=get_recipes_grouped(),
        products=get_products(location_id=location_id)
    )

@app.route('/audit-logs')
def audit_logs():

    if not login_required():
        return redirect(url_for('login'))

    return render_template(
        'audit_logs.html',
        logs=get_audit_logs(current_location_id())
    )


@app.route('/inventory', methods=['GET', 'POST'])
def inventory():

    if not login_required():
        return redirect(url_for('login'))

    location_id = current_location_id()
    products = get_stock_balances(location_id)

    if request.method == 'POST':

        if not admin_required():
            return redirect(url_for('inventory'))

        check_id = create_inventory_check(
            location_id,
            session.get('user')
        )

        for product in products:

            actual_quantity = request.form.get(
                f'product_{product["id"]}'
            )

            if actual_quantity is None:
                continue

            system_quantity = product['quantity']

            add_inventory_item(
                check_id,
                product['id'],
                system_quantity,
                actual_quantity
            )

            difference = float(actual_quantity) - float(system_quantity)

            if difference > 0:

                add_stock(
                    product['id'],
                    difference,
                    'Корректировка по ревизии',
                    location_id
                )

            elif difference < 0:

                write_off_stock(
                    product['id'],
                    abs(difference),
                    'Корректировка по ревизии',
                    location_id
                )

        add_audit_log(
            session.get('user'),
            location_id,
            'Ревизия склада',
            'Инвентаризация',
            'Проведена ревизия и корректировка остатков'
        )

        flash('Ревизия сохранена', 'success')

        return redirect(url_for('movements'))

    return render_template(
        'inventory.html',
        products=products
    )


@app.route('/inventory-history')
def inventory_history():

    if not login_required():
        return redirect(url_for('login'))

    return render_template(
        'inventory_history.html',
        checks=get_inventory_history(current_location_id())
    )


@app.route('/inventory/<int:check_id>')
def inventory_detail(check_id):

    if not login_required():
        return redirect(url_for('login'))

    return render_template(
        'inventory_detail.html',
        items=get_inventory_items(check_id)
    )


if __name__ == '__main__':
    app.run(debug=True)
